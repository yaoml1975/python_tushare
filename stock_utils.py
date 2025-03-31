# filename: stock_utils.py
import datetime
import inspect
import logging
import os
import time

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from data_cache import dc


def setup_logger(name=None):
    """
    初始化日志配置，默认日志文件名与调用此函数的脚本文件同名

    :param name: 指定日志文件名（不带扩展名），默认为调用此函数的脚本名
    :return: 返回配置好的 logger 实例
    """
    if name is None:
        # 获取调用该函数的文件名
        caller_file = inspect.stack()[1].filename
        name = os.path.splitext(os.path.basename(caller_file))[0]

    log_file = os.path.join(dc.log_dir, f"{name}.log")

    logger = logging.getLogger(name)
    if not logger.hasHandlers():  # 防止重复添加 handler
        logger.setLevel(logging.INFO)

        # formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
        formatter = logging.Formatter(
            '[%(asctime)s] [%(module)s:%(funcName)s] [%(levelname)s]- %(message)s')

        # 控制台日志
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)

        # 文件日志
        file_handler = logging.FileHandler(log_file, mode="a", encoding="utf-8")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

    return logger


logger = setup_logger()


def get_formatted_time(fmt: str = "%m%d%H%M%S") -> str:
    """
    获取当前时间的字符串表示，支持格式化。

    :param fmt: 时间格式，默认为 "%Y%m%d%H%M%S"（年日月时分秒）
    :return: 格式化后的时间字符串
    """
    return datetime.datetime.now().strftime(fmt)


def is_file_older_than(file_path, days=1):
    """判断文件是否早于指定天数"""
    if not os.path.exists(file_path):
        return False
    modification_time = os.path.getmtime(file_path)
    threshold_time = time.time() - (days * 86400)
    return modification_time <= threshold_time


def get_quarter_end_dates(year):
    """
    获取指定年份的季度结束日期
    """
    quarter_end_months = [3, 6, 9, 12]
    quarter_names = ['Q1', 'Q2', 'Q3', 'Q4']
    end_dates = {}
    for i, month in enumerate(quarter_end_months):
        last_day = datetime.date(year, month, 1)
        next_month = month % 12 + 1
        if next_month == 1:
            last_day = last_day.replace(year=year + 1)
        quarter_end = (datetime.date(last_day.year, next_month, 1) - datetime.timedelta(days=1))
        end_dates[quarter_names[i]] = quarter_end.strftime('%Y%m%d')
    return end_dates  # 这里增加了返回语句


def generate_quarter_list(start_year):
    """
    生成从指定年份到当前年份的季度列表
    """
    current_date = datetime.date.today()
    current_year, current_month = current_date.year, current_date.month
    quarter_map = {1: 'Q1', 2: 'Q1', 3: 'Q1', 4: 'Q2', 5: 'Q2', 6: 'Q2',
                   7: 'Q3', 8: 'Q3', 9: 'Q3', 10: 'Q4', 11: 'Q4', 12: 'Q4'}
    current_quarter = quarter_map[current_month]
    quarter_list = []
    for year in range(start_year, current_year + 1):
        quarter_dates = get_quarter_end_dates(year)
        # print(f"获取到的季度结束日期：{quarter_dates}")  # 打印季度结束日期
        for quarter, date in quarter_dates.items():
            if year == current_year and quarter == current_quarter:
                return quarter_list  # 不包含当前季度
            # quarter_list.append(f"{year}{quarter}:{date}")
            quarter_list.append(f"{date}")
    return quarter_list


def get_display_width(value):
    """计算字符串的显示宽度，ASCII字符计1，其他字符（如中文）计2"""
    if value is None:
        return 0
    value = str(value)
    return sum(2 if ord(c) > 127 else 1 for c in value)


def auto_adjust_column_width(file_path):
    """自动调整 Excel 列宽以适应内容，考虑中文字符"""
    wb = load_workbook(file_path)
    ws = wb.active

    max_widths = {}

    # 遍历所有单元格，记录每列的最大显示宽度
    for row in ws.iter_rows():
        for cell in row:
            col_letter = cell.column_letter
            current_width = get_display_width(cell.value)
            if col_letter not in max_widths or current_width > max_widths[col_letter]:
                max_widths[col_letter] = current_width

    # 处理可能存在的空列（根据最大列号）
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter not in max_widths:
            max_widths[col_letter] = 0  # 全空列宽度设为0

    # 调整列宽，添加边距并限制最大宽度
    for col_letter, max_length in max_widths.items():
        adjusted_width = max_length + 2  # 增加边距
        adjusted_width = min(adjusted_width, 50)  # 限制最大宽度为50
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(file_path)
    print(f"列宽已调整并保存到文件: {file_path}")


def get_last_trade_date():
    """获取最近一个交易日"""
    try:
        today = datetime.datetime.today()
        current_time = today.time()
        today_str = today.strftime('%Y%m%d')
        df = dc.pro.trade_cal(start_date=(today - datetime.timedelta(days=30)).strftime('%Y%m%d'),
                              end_date=today_str)
        if not df.empty:
            trade_days = sorted(df[df['is_open'] == 1]['cal_date'].tolist(), reverse=True)
            if today_str in trade_days:
                if current_time >= datetime.time(17, 0):
                    return today_str
                trade_days.remove(today_str)
            return trade_days[0] if trade_days else None
        logger.info("未找到有效的交易日")
        return None
    except Exception as e:
        logger.error(f"获取最近交易日失败：{e}")
        return None


def get_friday_trade_dates(year, month):
    """
    获取指定年和月中所有为星期五的交易日（过滤掉超过当前日期的交易日）

    参数:
      year: int，年份，例如2025
      month: int，月份，例如2

    返回:
      List[str]，符合条件的交易日列表，日期格式为 'YYYYMMDD'
    """
    try:
        # 构造当月的起始日期和结束日期
        start_date = datetime.date(year, month, 1)
        if month == 12:
            end_date = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            end_date = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
        start_str = start_date.strftime('%Y%m%d')
        end_str = end_date.strftime('%Y%m%d')

        # 获取该时间段内的交易日历数据
        df = dc.pro.trade_cal(start_date=start_str, end_date=end_str)
        if not df.empty:
            # 过滤出交易日（is_open==1）
            df_open = df[df['is_open'] == 1]
            friday_dates = []
            # 当前日期
            current_date = datetime.date.today()
            for date_str in df_open['cal_date']:
                # 将日期字符串转换为日期对象
                date_obj = datetime.datetime.strptime(date_str, '%Y%m%d').date()
                # 如果日期超过当前日期，则跳过
                if date_obj > current_date:
                    continue
                # weekday() 返回0-6，星期五为4
                if date_obj.weekday() == 4:
                    friday_dates.append(date_str)
            return friday_dates
        else:
            logger.info("未找到有效的交易日")
            return []
    except Exception as e:
        logger.error(f"获取交易日失败：{e}")
        return []


def get_last_n_trade_dates(n=20):
    """获取最近 n 个交易日"""
    try:
        today = datetime.datetime.today()
        current_time = today.time()
        today_str = today.strftime('%Y%m%d')
        df = dc.pro.trade_cal(start_date=(today - datetime.timedelta(days=2 * n)).strftime('%Y%m%d'),
                              end_date=today_str)
        if not df.empty:
            trade_days = sorted([date for date in df[df['is_open'] == 1]['cal_date'].tolist() if date <= today_str],
                                reverse=True)
            if today_str in trade_days and current_time < datetime.time(17, 0):
                trade_days.remove(today_str)
            return trade_days[:n]
        logger.info("未找到有效的交易日")
        return []
    except Exception as e:
        logger.error(f"获取最近 {n} 个交易日失败：{e}")
        return []


def fetch_stock_basic(trade_date, is_save_csv=True):
    """获取股票基础信息并保存为 CSV"""
    df = dc.pro.stock_basic(exchange='', list_status='L')
    if is_save_csv:
        filename = f"tushare_stock_basic_{trade_date}.csv"
        df.to_csv(os.path.join(dc.csv_dir, filename), index=False, encoding='utf-8_sig')
        logger.info(f"股票基础信息已保存至 {filename}")
    return df


def fetch_daily(trade_date, is_save_csv=True):
    """获取日线行情数据并保存为 CSV"""
    df = dc.pro.daily(trade_date=trade_date)
    if is_save_csv:
        filename = f"tushare_daily_{trade_date}.csv"
        df.to_csv(os.path.join(dc.csv_dir, filename), index=False, encoding='utf-8_sig')
        logger.info(f"日线行情数据已保存至 {filename}")
    return df


def fetch_daily_basic(trade_date, is_save_csv=True):
    """获取每日指标数据并保存为 CSV"""
    df = dc.pro.daily_basic(trade_date=trade_date)
    if is_save_csv:
        filename = f"tushare_daily_basic_{trade_date}.csv"
        df.to_csv(os.path.join(dc.csv_dir, filename), index=False, encoding='utf-8_sig')
        logger.info(f"每日指标数据已保存至 {filename}")
    return df


def fetch_weekly(trade_date, is_save_csv=True):
    """
    获取A股周线行情数据并保存为 CSV
    trade_date	str	N	交易日期 （每周最后一个交易日期，YYYYMMDD格式）
    """

    df = dc.pro.weekly(trade_date=trade_date)
    if is_save_csv:
        filename = f"tushare_weekly_{trade_date}.csv"
        full_path = os.path.join(dc.csv_dir, filename)
        df.to_csv(full_path, index=False, encoding='utf-8_sig')
        logger.info(f"周线行情数据已保存至 {filename}")
    return df


def fetch_stk_factor_pro_by_tscode(ts_code, start_date, end_date):
    """获取指定日期范围内的股票因子数据"""
    # 延时设置，避免超过访问频率限制
    time.sleep(0.5)  # 每次请求间隔0.3秒

    df = dc.pro.stk_factor_pro(ts_code=ts_code, start_date=start_date, end_date=end_date)
    return df.sort_values('trade_date').dropna()


def ensure_sufficient_data(ts_code, min_days, max_attempts=3601975):
    """确保获取足够天数的股票数据"""
    start_date = (datetime.datetime.today() - datetime.timedelta(days=2 * min_days)).strftime('%Y%m%d')
    end_date = datetime.datetime.today().strftime('%Y%m%d')

    attempt = 0
    while attempt < max_attempts:
        df = fetch_stk_factor_pro_by_tscode(ts_code, start_date, end_date)
        if len(df) >= min_days:
            return df
        # 如果数据不足，将 start_date 向前推移 min_days 天
        start_date = (datetime.datetime.strptime(start_date, '%Y%m%d') - datetime.timedelta(days=min_days)).strftime(
            '%Y%m%d')
        logger.info(f'{ts_code} 重新生成 start_date: {start_date}， 继续请求API...')
        attempt += 1

    # 如果达到最大尝试次数仍未获取足够数据，返回 None 或抛出异常
    logger.info(f'{ts_code} 无法获取足够数据，已达到最大尝试次数 {max_attempts}')
    return None


def get_recent_kdj_death_cross(ts_code):
    """获取最近一次 KDJ 死叉日期"""
    df_kdj = ensure_sufficient_data(ts_code, min_days=10)
    if df_kdj is None:
        logger.info(f"{ts_code}: 无法获取足够的数据来计算 KDJ 死叉")
        return get_last_trade_date()

    # 计算 KDJ 死叉
    df_kdj['kdj_death'] = (df_kdj['kdj_k_qfq'].shift(1) > df_kdj['kdj_d_qfq'].shift(1)) & (
            df_kdj['kdj_k_qfq'] < df_kdj['kdj_d_qfq'])

    if 'kdj_death' not in df_kdj.columns:
        raise ValueError("DataFrame 中缺少 'kdj_death' 列")

    # 获取最近一次死叉日期
    last_death = df_kdj[df_kdj['kdj_death']].tail(1)
    if not last_death.empty:
        return str(last_death['trade_date'].values[0]).replace('-', '')
    return None


def get_recent_macd_death_cross(ts_code):
    """获取最近一次 MACD 死叉日期"""
    df_macd = ensure_sufficient_data(ts_code, min_days=36)
    if df_macd is None:
        logger.info(f"{ts_code}: 无法获取足够的数据来计算 MACD 死叉")
        return get_last_trade_date()

    # 计算 MACD 死叉
    df_macd['macd_death'] = (df_macd['macd_dif_qfq'].shift(1) > df_macd['macd_dea_qfq'].shift(1)) & (
            df_macd['macd_dif_qfq'] < df_macd['macd_dea_qfq'])

    if 'macd_death' not in df_macd.columns:
        raise ValueError("DataFrame 中缺少 'macd_death' 列")

    # 获取最近一次死叉日期
    last_death = df_macd[df_macd['macd_death']].tail(1)
    if not last_death.empty:
        return str(last_death['trade_date'].values[0]).replace('-', '')
    return None


def fetch_fina_indicator_vip_by_tscode(ts_code, quarter_list, is_save_csv=True):
    # 存储ts_code所有财务数据的列表
    all_data = []

    for item in tqdm(quarter_list, total=len(quarter_list), desc=f"获取股票 {ts_code} 财务数据"):
        # 获取报告期的结束日期
        quarter_end_date = item.split(":")[1]
        logger.info(f"开始获取{ts_code} - {quarter_end_date} 的财务数据...")

        # 重试3次
        for _ in range(3):
            try:
                df = dc.pro.fina_indicator_vip(ts_code=ts_code, period=quarter_end_date,
                                               fields='ts_code,'
                                                      'ann_date,'
                                                      'end_date,'
                                                      'roe,'
                                                      'fcff,'
                                                      'grossprofit_margin,'
                                                      'equity_yoy,'
                                                      'debt_to_assets,'
                                                      'update_flag',
                                               update_flag='1')
                if not df.empty:
                    all_data.append(df)
                    print(f"成功获取{ts_code} - {quarter_end_date} 的数据。")
                    break
                else:
                    logger.warning(f"未获取到{ts_code} - {quarter_end_date} 的数据，继续重试...")

            except Exception as e:
                # 获取数据失败，重试
                logger.error(f"{ts_code} - [fina_indicator_vip] 获取财务指标失败，重试中: {str(e)}")
                time.sleep(5)

    # 如果没有获取到任何数据，提前返回
    if not all_data:
        print("未获取到任何数据，无法合并。")
        return

    # 在拼接前去除所有空列（全是NaN的列）
    # all_data = [df.dropna(axis=1, how='all') for df in all_data]

    # 合并所有数据
    final_data = pd.concat(all_data, ignore_index=True)

    if is_save_csv:
        filename = f"tushare_fina_indicator_vip_{ts_code}.csv"
        final_data.to_csv(os.path.join(dc.csv_dir, filename), index=False, encoding='utf-8-sig')
        logger.info(f"{ts_code} 的财务数据已保存至 {filename}")

    logger.info(final_data)

    return final_data


def fetch_fina_indicator_vip_by_tscode_quarter_str(ts_code, quarter_str, is_save_csv=False):
    # 存储ts_code所有财务数据的列表
    all_data = []

    # 重试3次
    for _ in range(3):
        try:
            df = dc.pro.fina_indicator_vip(ts_code=ts_code, period=quarter_str,
                                           fields='ts_code,'
                                                  'ann_date,'
                                                  'end_date,'
                                                  'roe,'
                                                  'fcff,'
                                                  'grossprofit_margin,'
                                                  'equity_yoy,'
                                                  'debt_to_assets,'
                                                  'update_flag',
                                           update_flag='1')
            if not df.empty:
                all_data.append(df)
                # logger.info(f"成功获取{ts_code} - {quarter_str} 的数据。")
                break
            else:
                logger.warning(f"未获取到{ts_code} - {quarter_str} 的数据，继续重试...")

        except Exception as e:
            # 获取数据失败，重试
            logger.error(f"{ts_code} - [fina_indicator_vip] 获取财务指标失败，重试中: {str(e)}")
            time.sleep(5)

    # 如果没有获取到任何数据，提前返回
    if not all_data:
        logger.warning("未获取到任何数据，无法合并。")
        return

    # 在拼接前去除所有空列（全是NaN的列）
    # all_data = [df.dropna(axis=1, how='all') for df in all_data]

    # 合并所有数据
    final_data = pd.concat(all_data, ignore_index=True)

    if is_save_csv:
        filename = f"tushare_fina_indicator_vip_{ts_code}.csv"
        final_data.to_csv(os.path.join(dc.csv_dir, filename), index=False, encoding='utf-8-sig')
        logger.info(f"{ts_code} 的财务数据已保存至 {filename}")

    # logger.info(final_data)

    return final_data


def fetch_fina_indicator_vip_by_quarter_str(quarter_str, is_save_csv=True):
    filename = f"tushare_fina_indicator_vip_{quarter_str}.csv"
    full_path = os.path.join(dc.csv_dir, filename)
    if os.path.exists(full_path):
        # logger.info(f"{filename} 已经存在。")
        df = pd.read_csv(full_path)
        return df

    # 存储ts_code所有财务数据的列表
    all_data = []

    # 重试3次
    for _ in range(3):
        try:
            logger.info(f"开始获取 {quarter_str} 的数据...")
            df = dc.pro.fina_indicator_vip(ts_code='', period=quarter_str,
                                           fields='ts_code,ann_date,end_date,eps,dt_eps,total_revenue_ps,revenue_ps,'
                                                  'capital_rese_ps,surplus_rese_ps,undist_profit_ps,extra_item,'
                                                  'profit_dedt,gross_margin,current_ratio,quick_ratio,cash_ratio,'
                                                  'invturn_days,arturn_days,inv_turn,ar_turn,ca_turn,fa_turn,'
                                                  'assets_turn,op_income,valuechange_income,interst_income,daa,'
                                                  'ebit,ebitda,fcff,fcfe,current_exint,noncurrent_exint,interestdebt,'
                                                  'netdebt,tangible_asset,working_capital,networking_capital,'
                                                  'invest_capital,retained_earnings,diluted2_eps,bps,ocfps,'
                                                  'retainedps,cfps,ebit_ps,fcff_ps,fcfe_ps,netprofit_margin,'
                                                  'grossprofit_margin,cogs_of_sales,expense_of_sales,profit_to_gr,'
                                                  'saleexp_to_gr,adminexp_of_gr,finaexp_of_gr,impai_ttm,gc_of_gr,'
                                                  'op_of_gr,ebit_of_gr,roe,roe_waa,roe_dt,roa,npta,roic,roe_yearly,'
                                                  'roa2_yearly,roe_avg,opincome_of_ebt,investincome_of_ebt,'
                                                  'n_op_profit_of_ebt,tax_to_ebt,dtprofit_to_profit,salescash_to_or,'
                                                  'ocf_to_or,ocf_to_opincome,capitalized_to_da,debt_to_assets,'
                                                  'assets_to_eqt,dp_assets_to_eqt,ca_to_assets,nca_to_assets,'
                                                  'tbassets_to_totalassets,int_to_talcap,eqt_to_talcapital,'
                                                  'currentdebt_to_debt,longdeb_to_debt,ocf_to_shortdebt,'
                                                  'debt_to_eqt,eqt_to_debt,eqt_to_interestdebt,tangibleasset_to_debt,'
                                                  'tangasset_to_intdebt,tangibleasset_to_netdebt,ocf_to_debt,'
                                                  'ocf_to_interestdebt,ocf_to_netdebt,ebit_to_interest,'
                                                  'longdebt_to_workingcapital,ebitda_to_debt,turn_days,'
                                                  'roa_yearly,roa_dp,fixed_assets,profit_prefin_exp,non_op_profit,'
                                                  'op_to_ebt,nop_to_ebt,ocf_to_profit,cash_to_liqdebt,'
                                                  'cash_to_liqdebt_withinterest,op_to_liqdebt,op_to_debt,roic_yearly,'
                                                  'total_fa_trun,profit_to_op,q_opincome,q_investincome,q_dtprofit,'
                                                  'q_eps,q_netprofit_margin,q_gsprofit_margin,q_exp_to_sales,'
                                                  'q_profit_to_gr,q_saleexp_to_gr,q_adminexp_to_gr,q_finaexp_to_gr,'
                                                  'q_impair_to_gr_ttm,q_gc_to_gr,q_op_to_gr,q_roe,q_dt_roe,q_npta,'
                                                  'q_opincome_to_ebt,q_investincome_to_ebt,q_dtprofit_to_profit,'
                                                  'q_salescash_to_or,q_ocf_to_sales,q_ocf_to_or,basic_eps_yoy,'
                                                  'dt_eps_yoy,cfps_yoy,op_yoy,ebt_yoy,netprofit_yoy,dt_netprofit_yoy,'
                                                  'ocf_yoy,roe_yoy,bps_yoy,assets_yoy,eqt_yoy,tr_yoy,or_yoy,'
                                                  'q_gr_yoy,q_gr_qoq,q_sales_yoy,q_sales_qoq,q_op_yoy,q_op_qoq,'
                                                  'q_profit_yoy,q_profit_qoq,q_netprofit_yoy,q_netprofit_qoq,'
                                                  'equity_yoy,rd_exp,update_flag',
                                           update_flag='1')
            if not df.empty:
                all_data.append(df)
                logger.info(f"成功获取 {quarter_str} 的数据。")
                break
            else:
                logger.warning(f"未获取到 {quarter_str} 的数据，继续重试...")

        except Exception as e:
            # 获取数据失败，重试
            logger.error(f"获取 {quarter_str} 财务指标失败，重试中: {str(e)}")
            time.sleep(5)

    # 如果没有获取到任何数据，提前返回
    if not all_data:
        logger.warning("未获取到任何数据，无法合并。")
        return None

    # 在拼接前去除所有空列（全是NaN的列）
    all_data = [df.dropna(axis=1, how='all') for df in all_data]

    # 合并所有数据
    final_data = pd.concat(all_data, ignore_index=True)

    if is_save_csv:
        final_data.to_csv(full_path, index=False, encoding='utf-8-sig')
        logger.info(f"{quarter_str} 的财务数据已保存至 {full_path}")

    # logger.info(final_data)

    return final_data


def load_csv(file_name, fetch_function, trade_date):
    """加载指定交易日的 CSV 文件"""
    if 'filter' in file_name:
        file_path = os.path.join(dc.filter_dir, f"{file_name}_{trade_date}.csv")
    else:
        file_path = os.path.join(dc.csv_dir, f"{file_name}_{trade_date}.csv")
    if not os.path.exists(file_path):
        logger.error(f"加载指定交易日的 CSV 文件不存在: {file_path}, 重新生成...")
        fetch_function(trade_date)
    return pd.read_csv(file_path)


if __name__ == '__main__':
    logger.info(get_quarter_end_dates(2025))
    logger.info('-' * 100)

    logger.info(generate_quarter_list(2024))
    logger.info('-' * 100)

    # 循环读取并打印每个季度的结束日期
    # all_quarter_list = generate_quarter_list(2024)
    # for item in all_quarter_list:
    #     quarter_end_date = item.split(":")[1]  # 根据格式拆分，提取日期部分
    #     logger.info(quarter_end_date)

    # logger.info(get_last_trade_date())
    # logger.info('-' * 100)

    logger.info(get_last_n_trade_dates())
    logger.info('-' * 100)

    logger.info(get_last_n_trade_dates(n=20))
    logger.info('-' * 100)

    logger.info(get_last_n_trade_dates(n=100))
    logger.info('-' * 100)

    ts_code = '000001.SZ'
    # **筛选 2023 Q4**
    quarter = get_quarter_end_dates(2024)['Q3']
    logger.info(fetch_fina_indicator_vip_by_tscode_quarter_str('ts_code', quarter))
